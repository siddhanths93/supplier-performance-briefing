from io import BytesIO

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


def create_supplier_briefing_export(
    supplier_findings: pd.DataFrame,
) -> pd.DataFrame:
    """
    Create a clean supplier findings export table.
    """
    if supplier_findings.empty:
        return pd.DataFrame()

    export_data = supplier_findings.copy()

    export_data = export_data[
        [
            "supplier_name",
            "category",
            "annual_spend",
            "supplier_attention_score",
            "attention_level",
            "data_confidence_pct",
            "observation",
            "implication",
            "suggested_next_step",
        ]
    ]

    export_data = export_data.rename(
        columns={
            "supplier_name": "Supplier",
            "category": "Category",
            "annual_spend": "Annual Spend",
            "supplier_attention_score": "Attention Score",
            "attention_level": "Attention Level",
            "data_confidence_pct": "Data Confidence %",
            "observation": "Observation",
            "implication": "Implication",
            "suggested_next_step": "Suggested Next Step",
        }
    )

    return export_data


def create_category_briefing_export(
    category_findings: pd.DataFrame,
) -> pd.DataFrame:
    """
    Create a clean category findings export table.
    """
    if category_findings.empty:
        return pd.DataFrame()

    export_data = category_findings.copy()

    export_data = export_data[
        [
            "category",
            "total_category_spend",
            "observation",
            "implication",
            "suggested_next_step",
        ]
    ]

    export_data = export_data.rename(
        columns={
            "category": "Category",
            "total_category_spend": "Total Category Spend",
            "observation": "Observation",
            "implication": "Implication",
            "suggested_next_step": "Suggested Next Step",
        }
    )

    return export_data


def create_top_supplier_scores_export(
    supplier_data: pd.DataFrame,
    top_n: int = 25,
) -> pd.DataFrame:
    """
    Create a clean export of the top supplier scores.
    """
    if supplier_data.empty:
        return pd.DataFrame()

    score_columns = [
        "supplier_name",
        "category",
        "annual_spend",
        "category_spend_share_pct",
        "spend_exposure_score",
        "delivery_risk_score",
        "quality_risk_score",
        "criticality_score",
        "supplier_attention_score",
        "attention_level",
        "data_confidence_pct",
    ]

    export_data = (
        supplier_data[score_columns]
        .sort_values(
            by="supplier_attention_score",
            ascending=False,
        )
        .head(top_n)
        .copy()
    )

    export_data = export_data.rename(
        columns={
            "supplier_name": "Supplier",
            "category": "Category",
            "annual_spend": "Annual Spend",
            "category_spend_share_pct": "Category Spend Share %",
            "spend_exposure_score": "Spend Exposure Score",
            "delivery_risk_score": "Delivery Risk Score",
            "quality_risk_score": "Quality Risk Score",
            "criticality_score": "Criticality Score",
            "supplier_attention_score": "Attention Score",
            "attention_level": "Attention Level",
            "data_confidence_pct": "Data Confidence %",
        }
    )

    return export_data


def create_category_metrics_export(
    category_metrics: pd.DataFrame,
) -> pd.DataFrame:
    """
    Create a clean category metrics export.
    """
    if category_metrics.empty:
        return pd.DataFrame()

    export_data = category_metrics.copy()

    export_data = export_data.rename(
        columns={
            "category": "Category",
            "total_category_spend": "Total Category Spend",
            "supplier_count": "Supplier Count",
            "average_spend_per_supplier": "Average Spend per Supplier",
            "median_spend_per_supplier": "Median Spend per Supplier",
            "top_supplier_share_pct": "Top Supplier Share %",
            "top_3_supplier_share_pct": "Top 3 Supplier Share %",
            "top_5_supplier_share_pct": "Top 5 Supplier Share %",
            "suppliers_to_80_pct_spend": "Suppliers to 80% Spend",
            "hhi": "HHI",
            "tail_supplier_count": "Tail Supplier Count",
            "tail_supplier_pct": "Tail Supplier %",
            "tail_spend": "Tail Spend",
            "tail_spend_pct": "Tail Spend %",
            "concentration_risk_flag": "Concentration Review Flag",
            "fragmentation_review_flag": "Fragmentation Review Flag",
            "tail_spend_review_flag": "Tail Spend Review Flag",
        }
    )

    return export_data


def create_methodology_export() -> pd.DataFrame:
    """
    Create a methodology tab for the briefing workbook.
    """
    methodology_rows = [
        {
            "Section": "Purpose",
            "Description": (
                "This workbook summarizes supplier and category "
                "relationships that may deserve management review."
            ),
        },
        {
            "Section": "Data source",
            "Description": (
                "The prototype is designed for supplier-level spend "
                "and performance data. The sample dataset is synthetic."
            ),
        },
        {
            "Section": "Supplier attention score",
            "Description": (
                "The illustrative score combines spend exposure, "
                "delivery deterioration, quality deterioration, and "
                "supplier criticality."
            ),
        },
        {
            "Section": "Score weights",
            "Description": (
                "Spend exposure: 30%; delivery risk: 25%; "
                "quality risk: 25%; criticality: 20%."
            ),
        },
        {
            "Section": "Data confidence",
            "Description": (
                "Data confidence reflects the percentage of scoring "
                "inputs available for a supplier. Missing evidence "
                "reduces confidence rather than automatically increasing risk."
            ),
        },
        {
            "Section": "Category concentration",
            "Description": (
                "Category concentration is evaluated using top-supplier "
                "share, top-three supplier share, suppliers required to "
                "reach 80% of spend, and HHI."
            ),
        },
        {
            "Section": "Important limitation",
            "Description": (
                "Findings are review hypotheses, not final sourcing "
                "decisions, guaranteed savings estimates, or supplier "
                "performance judgments."
            ),
        },
    ]

    return pd.DataFrame(methodology_rows)


def format_excel_worksheet(
        worksheet,
        dataframe: pd.DataFrame,
) -> None:
    """
    Apply basic professional formatting to an Excel worksheet.
    """
    header_fill = PatternFill(
        fill_type="solid",
        fgColor="1F4E78",
    )

    header_font = Font(
        bold=True,
        color="FFFFFF",
    )

    header_alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True,
    )

    body_alignment = Alignment(
        vertical="top",
        wrap_text=True,
    )

    worksheet.freeze_panes = "A2"

    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    for row in worksheet.iter_rows(
            min_row=2,
            max_row=worksheet.max_row,
    ):
        for cell in row:
            cell.alignment = body_alignment

    for column_index, column_name in enumerate(
            dataframe.columns,
            start=1,
    ):
        column_letter = get_column_letter(
            column_index
        )

        if column_name in [
            "Observation",
            "Implication",
            "Suggested Next Step",
            "Description",
        ]:
            worksheet.column_dimensions[
                column_letter
            ].width = 55

        elif "%" in column_name:
            worksheet.column_dimensions[
                column_letter
            ].width = 18

        elif "Share" in column_name:
            worksheet.column_dimensions[
                column_letter
            ].width = 18

        elif "Score" in column_name:
            worksheet.column_dimensions[
                column_letter
            ].width = 18

        elif "Spend" in column_name:
            worksheet.column_dimensions[
                column_letter
            ].width = 18

        else:
            worksheet.column_dimensions[
                column_letter
            ].width = 22

    for row in worksheet.iter_rows(
            min_row=2,
            max_row=worksheet.max_row,
    ):
        for cell in row:
            header_value = worksheet.cell(
                row=1,
                column=cell.column,
            ).value

            if header_value is None:
                continue

            header_text = str(header_value)

            if "%" in header_text:
                cell.number_format = '0.0'

            elif "Share" in header_text:
                cell.number_format = '0.0'

            elif "Score" in header_text:
                cell.number_format = '0.0'

            elif "Spend" in header_text:
                cell.number_format = '$#,##0'


def create_excel_briefing_workbook(
    supplier_findings: pd.DataFrame,
    category_findings: pd.DataFrame,
    supplier_data: pd.DataFrame,
    category_metrics: pd.DataFrame,
) -> bytes:
    """
    Create a multi-tab Excel briefing workbook in memory.
    """
    output = BytesIO()

    supplier_findings_export = (
        create_supplier_briefing_export(
            supplier_findings
        )
    )

    category_findings_export = (
        create_category_briefing_export(
            category_findings
        )
    )

    top_supplier_scores_export = (
        create_top_supplier_scores_export(
            supplier_data
        )
    )

    category_metrics_export = (
        create_category_metrics_export(
            category_metrics
        )
    )

    methodology_export = create_methodology_export()

    with pd.ExcelWriter(
        output,
        engine="openpyxl",
    ) as writer:
        supplier_findings_export.to_excel(
            writer,
            sheet_name="Supplier Findings",
            index=False,
        )

        category_findings_export.to_excel(
            writer,
            sheet_name="Category Findings",
            index=False,
        )

        top_supplier_scores_export.to_excel(
            writer,
            sheet_name="Top 25 Supplier Scores",
            index=False,
        )

        category_metrics_export.to_excel(
            writer,
            sheet_name="Category Metrics",
            index=False,
        )

        methodology_export.to_excel(
            writer,
            sheet_name="Methodology",
            index=False,
        )

        workbook = writer.book

        worksheet_dataframes = {
            "Supplier Findings": supplier_findings_export,
            "Category Findings": category_findings_export,
            "Top 25 Supplier Scores": top_supplier_scores_export,
            "Category Metrics": category_metrics_export,
            "Methodology": methodology_export,
        }

        for sheet_name, dataframe in (
            worksheet_dataframes.items()
        ):
            worksheet = workbook[sheet_name]

            format_excel_worksheet(
                worksheet,
                dataframe,
            )

    output.seek(0)

    return output.getvalue()